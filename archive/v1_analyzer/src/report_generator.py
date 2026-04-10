"""レポート生成モジュール（PDF + Excel）"""

import datetime
from pathlib import Path
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, HRFlowable
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# フォント登録
_FONT_REGISTERED = False

def _register_fonts():
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return
    # macOS日本語フォント
    font_paths = [
        ("/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc", "HiraKaku"),
        ("/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc", "HiraKakuBold"),
        ("/System/Library/Fonts/Hiragino Sans GB.ttc", "HiraKaku"),
        ("/Library/Fonts/Arial Unicode.ttf", "ArialUnicode"),
    ]
    for path, name in font_paths:
        try:
            pdfmetrics.registerFont(TTFont(name, path, subfontIndex=0))
            _FONT_REGISTERED = True
            return
        except Exception:
            continue

    # フォールバック: Helvetica（日本語不可だが動作はする）
    _FONT_REGISTERED = True


def _get_font_name():
    _register_fonts()
    for name in ["HiraKaku", "ArialUnicode"]:
        try:
            pdfmetrics.getFont(name)
            return name
        except Exception:
            continue
    return "Helvetica"


def _fmt(value, fmt_type="number"):
    """数値フォーマット"""
    if value is None:
        return "N/A"
    if fmt_type == "yen":
        return f"¥{value:,.0f}"
    elif fmt_type == "man":
        return f"{value / 10000:,.0f}万円"
    elif fmt_type == "percent":
        return f"{value:.1f}%"
    elif fmt_type == "number":
        return f"{value:,.0f}"
    return str(value)


def generate_pdf(
    property_data: dict,
    rental_result,
    minpaku_result,
    tax_comparison: dict,
    loan_results: list,
    guarantee_results: list,
    risks: list,
    exit_scenarios: list,
    sensitivity: dict,
    decision: dict,
    output_path: Path,
):
    """PDFレポート生成"""
    font_name = _get_font_name()

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()

    # カスタムスタイル
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        fontName=font_name, fontSize=18, spaceAfter=6 * mm,
    )
    h1_style = ParagraphStyle(
        "CustomH1", parent=styles["Heading1"],
        fontName=font_name, fontSize=14, spaceAfter=4 * mm, spaceBefore=6 * mm,
        textColor=colors.HexColor("#1a365d"),
    )
    h2_style = ParagraphStyle(
        "CustomH2", parent=styles["Heading2"],
        fontName=font_name, fontSize=11, spaceAfter=3 * mm, spaceBefore=4 * mm,
        textColor=colors.HexColor("#2c5282"),
    )
    body_style = ParagraphStyle(
        "CustomBody", parent=styles["Normal"],
        fontName=font_name, fontSize=9, leading=14,
    )
    small_style = ParagraphStyle(
        "Small", parent=styles["Normal"],
        fontName=font_name, fontSize=7, leading=10,
    )

    elements = []

    # === タイトル ===
    elements.append(Paragraph("不動産投資分析レポート", title_style))
    elements.append(Paragraph(
        f"物件名: {property_data.get('property_name', '不明')} | "
        f"生成日: {datetime.datetime.now().strftime('%Y/%m/%d')}",
        body_style
    ))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2c5282")))
    elements.append(Spacer(1, 4 * mm))

    # === 1. エグゼクティブサマリー ===
    elements.append(Paragraph("1. エグゼクティブサマリー（投資判断）", h1_style))

    verdict_color = {
        "買い推奨": "#38a169",
        "条件付き推奨": "#d69e2e",
        "慎重検討": "#e53e3e",
        "見送り推奨": "#c53030",
    }.get(decision["verdict"], "var(--border-secondary)")

    elements.append(Paragraph(
        f'<font color="{verdict_color}" size="14"><b>【{decision["verdict"]}】スコア: {decision["score"]}/100</b></font>',
        body_style
    ))
    elements.append(Paragraph(decision["verdict_detail"], body_style))
    elements.append(Spacer(1, 2 * mm))

    summary_data = [
        ["指標", "値", "評価基準"],
        ["実質利回り", _fmt(decision["key_metrics"]["net_yield"], "percent"), "4%以上が目安"],
        ["DSCR", f'{decision["key_metrics"]["dscr"]:.2f}', "1.2以上が安全圏"],
        ["CCR", _fmt(decision["key_metrics"]["cash_on_cash"], "percent"), "4%以上が目安"],
        ["IRR", _fmt(decision["key_metrics"]["irr"], "percent") if decision["key_metrics"]["irr"] else "N/A", "5%以上が望ましい"],
        ["損益分岐稼働率", _fmt(decision["key_metrics"]["breakeven_occupancy"], "percent"), "80%以下が安全"],
        ["高リスク項目数", str(decision["high_risk_count"]), "0が理想"],
    ]
    elements.append(_make_table(summary_data, [120, 80, 160]))

    # === 2. 物件概要 ===
    elements.append(Paragraph("2. 物件概要", h1_style))
    prop_info = [
        ["項目", "内容"],
        ["物件名", str(property_data.get("property_name", "N/A"))],
        ["所在地", str(property_data.get("address", "N/A"))],
        ["価格", _fmt(property_data.get("price", 0) * 10000, "man")],
        ["構造", str(property_data.get("structure", "N/A"))],
        ["築年", f'{property_data.get("year_built", "N/A")}年'],
        ["総戸数", str(property_data.get("total_units", "N/A"))],
        ["土地面積", f'{property_data.get("land_area_sqm", "N/A")}㎡'],
        ["建物面積", f'{property_data.get("building_area_sqm", "N/A")}㎡'],
        ["最寄り駅", f'{property_data.get("station", "N/A")} 徒歩{property_data.get("walk_minutes", "N/A")}分'],
        ["用途地域", str(property_data.get("zoning", "N/A"))],
        ["土地権利", str(property_data.get("land_rights", "N/A"))],
        ["表面利回り", _fmt(property_data.get("gross_yield"), "percent")],
    ]
    elements.append(_make_table(prop_info, [100, 260]))

    # === 3. 賃貸vs民泊比較 ===
    elements.append(PageBreak())
    elements.append(Paragraph("3. 賃貸 vs 民泊 収益比較", h1_style))
    compare = [
        ["項目", "賃貸", "民泊"],
        ["年間総収入", _fmt(rental_result.annual_gross_income, "yen"),
         _fmt(minpaku_result.annual_gross_income, "yen") if minpaku_result else "N/A"],
        ["年間経費", _fmt(rental_result.annual_operating_expenses, "yen"),
         _fmt(minpaku_result.annual_operating_expenses, "yen") if minpaku_result else "N/A"],
        ["NOI", _fmt(rental_result.annual_noi, "yen"),
         _fmt(minpaku_result.annual_noi, "yen") if minpaku_result else "N/A"],
        ["ローン返済", _fmt(rental_result.annual_debt_service, "yen"),
         _fmt(minpaku_result.annual_debt_service, "yen") if minpaku_result else "N/A"],
        ["税引前CF", _fmt(rental_result.annual_net_cf, "yen"),
         _fmt(minpaku_result.annual_net_cf, "yen") if minpaku_result else "N/A"],
        ["表面利回り", _fmt(rental_result.gross_yield, "percent"),
         _fmt(minpaku_result.gross_yield, "percent") if minpaku_result else "N/A"],
        ["実質利回り", _fmt(rental_result.net_yield, "percent"),
         _fmt(minpaku_result.net_yield, "percent") if minpaku_result else "N/A"],
        ["DSCR", f"{rental_result.dscr:.2f}",
         f"{minpaku_result.dscr:.2f}" if minpaku_result else "N/A"],
        ["CCR", _fmt(rental_result.cash_on_cash, "percent"),
         _fmt(minpaku_result.cash_on_cash, "percent") if minpaku_result else "N/A"],
        ["IRR", _fmt(rental_result.irr, "percent"),
         _fmt(minpaku_result.irr, "percent") if minpaku_result else "N/A"],
    ]
    elements.append(_make_table(compare, [100, 130, 130]))

    # === 4. 個人vs法人比較 ===
    elements.append(Paragraph("4. 個人 vs 法人 税務比較", h1_style))
    ind = tax_comparison["individual"]
    corp = tax_comparison["corporate"]
    tax_data = [
        ["項目", "個人", "法人"],
        ["課税所得", _fmt(ind.taxable_income, "yen"), _fmt(corp.taxable_income, "yen")],
        ["税額合計", _fmt(ind.total_tax, "yen"), _fmt(corp.total_tax, "yen")],
        ["実効税率", _fmt(ind.effective_tax_rate, "percent"), _fmt(corp.effective_tax_rate, "percent")],
        ["税引後手取り", _fmt(ind.net_income_after_tax, "yen"), _fmt(corp.net_income_after_tax, "yen")],
        ["10年累計手取り", _fmt(ind.ten_year_net_income, "yen"), _fmt(corp.ten_year_net_income, "yen")],
        ["初期費用", _fmt(ind.setup_cost, "yen"), _fmt(corp.setup_cost, "yen")],
        ["年間固定費", _fmt(ind.annual_overhead, "yen"), _fmt(corp.annual_overhead, "yen")],
    ]
    elements.append(_make_table(tax_data, [100, 130, 130]))
    elements.append(Paragraph(
        f'<b>推奨: {tax_comparison["recommendation"]}</b> - {tax_comparison["reason"]}',
        body_style
    ))

    # === 5. 融資戦略 ===
    elements.append(PageBreak())
    elements.append(Paragraph("5. 融資戦略（金融機関比較）", h1_style))
    loan_data = [["金融機関", "金利", "融資額", "月額返済", "年間返済", "可否", "推奨"]]
    for lr in loan_results:
        loan_data.append([
            lr.bank_name[:8],
            f"{lr.interest_rate}%",
            _fmt(lr.loan_amount, "man") if lr.available else "-",
            _fmt(lr.monthly_payment, "yen") if lr.available else "-",
            _fmt(lr.annual_payment, "yen") if lr.available else "-",
            "○" if lr.available else f"× {lr.reason}",
            "★" if lr.recommended else "",
        ])
    elements.append(_make_table(loan_data, [65, 35, 60, 65, 65, 85, 25], font_size=7))

    # 保証会社
    elements.append(Paragraph("保証会社推奨", h2_style))
    gc_data = [["会社名", "保証料率", "保証料概算", "マッチ度", "理由"]]
    for gc in guarantee_results:
        gc_data.append([
            gc.company_name,
            f"{gc.guarantee_fee_rate:.0%}",
            _fmt(gc.guarantee_fee_amount, "yen"),
            "★" * gc.match_score,
            gc.reason[:30],
        ])
    elements.append(_make_table(gc_data, [75, 45, 65, 55, 160], font_size=7))

    # === 6. リスク分析 ===
    elements.append(Paragraph("6. リスク分析", h1_style))
    risk_data = [["分類", "重要度", "リスク内容", "対策"]]
    for r in risks:
        risk_data.append([r.category, r.severity, r.description[:40], r.mitigation[:40]])
    elements.append(_make_table(risk_data, [40, 40, 160, 160], font_size=7))

    # === 7. 出口戦略 ===
    elements.append(PageBreak())
    elements.append(Paragraph("7. 出口戦略（売却シミュレーション）", h1_style))
    exit_data = [["売却年", "物件価値", "ローン残高", "累計CF", "売却手取り", "総利益", "年率ROI"]]
    for es in exit_scenarios:
        exit_data.append([
            f"{es.year}年目",
            _fmt(es.property_value, "man"),
            _fmt(es.loan_balance, "man"),
            _fmt(es.cumulative_cf, "man"),
            _fmt(es.sale_proceeds, "man"),
            _fmt(es.total_profit, "man"),
            _fmt(es.annualized_roi, "percent"),
        ])
    elements.append(_make_table(exit_data, [40, 65, 65, 55, 65, 60, 45], font_size=7))

    # === 8. 感度分析 ===
    elements.append(Paragraph("8. 感度分析", h1_style))

    for label, key in [("家賃変動", "rent"), ("稼働率変動", "occupancy"), ("金利変動", "interest")]:
        elements.append(Paragraph(label, h2_style))
        sens_data = [["変動", "実質利回り", "CCR", "DSCR", "年間CF"]]
        for delta_label, vals in sensitivity[key].items():
            sens_data.append([
                delta_label,
                _fmt(vals["net_yield"], "percent"),
                _fmt(vals["cash_on_cash"], "percent"),
                f'{vals["dscr"]:.2f}',
                _fmt(vals["annual_net_cf"], "yen"),
            ])
        elements.append(_make_table(sens_data, [60, 70, 70, 60, 100], font_size=7))

    # === フッター ===
    elements.append(Spacer(1, 10 * mm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    elements.append(Paragraph(
        f"本レポートは自動生成されたものであり、投資助言ではありません。投資判断は自己責任でお願いします。"
        f" | 生成日時: {datetime.datetime.now().strftime('%Y/%m/%d %H:%M')}",
        small_style,
    ))

    doc.build(elements)


def _make_table(data, col_widths, font_size=8):
    """テーブル生成ヘルパー"""
    font_name = _get_font_name()
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c5282")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("FONTSIZE", (0, 0), (-1, 0), font_size),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7fafc")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    t = Table(data, colWidths=col_widths)
    t.setStyle(TableStyle(style_cmds))
    return t


def generate_excel(
    property_data: dict,
    rental_result,
    minpaku_result,
    tax_comparison: dict,
    loan_results: list,
    risks: list,
    exit_scenarios: list,
    sensitivity: dict,
    decision: dict,
    output_path: Path,
):
    """Excel詳細レポート生成"""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_header(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # --- サマリーシート ---
    ws = wb.active
    ws.title = "サマリー"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    ws["A1"] = "不動産投資分析レポート"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"物件名: {property_data.get('property_name', 'N/A')}"
    ws["A3"] = f"生成日: {datetime.datetime.now().strftime('%Y/%m/%d')}"
    ws["A5"] = f"投資判断: {decision['verdict']}（スコア: {decision['score']}/100）"
    ws["A5"].font = Font(bold=True, size=12, color="FF0000" if "見送り" in decision["verdict"] else "008000")

    row = 7
    write_header(ws, row, ["指標", "賃貸", "民泊"])
    metrics = [
        ("表面利回り", rental_result.gross_yield, minpaku_result.gross_yield if minpaku_result else None),
        ("実質利回り", rental_result.net_yield, minpaku_result.net_yield if minpaku_result else None),
        ("DSCR", rental_result.dscr, minpaku_result.dscr if minpaku_result else None),
        ("CCR", rental_result.cash_on_cash, minpaku_result.cash_on_cash if minpaku_result else None),
        ("IRR", rental_result.irr, minpaku_result.irr if minpaku_result else None),
        ("年間NOI", rental_result.annual_noi, minpaku_result.annual_noi if minpaku_result else None),
        ("年間CF", rental_result.annual_net_cf, minpaku_result.annual_net_cf if minpaku_result else None),
    ]
    for i, (name, rental_val, minpaku_val) in enumerate(metrics, row + 1):
        ws.cell(row=i, column=1, value=name).border = thin_border
        ws.cell(row=i, column=2, value=rental_val).border = thin_border
        ws.cell(row=i, column=3, value=minpaku_val).border = thin_border

    # --- 35年CFシート ---
    ws_cf = wb.create_sheet("35年キャッシュフロー")
    headers = ["年", "総収入", "経費", "NOI", "ローン返済", "税引前CF", "累計CF", "ローン残高", "物件価値"]
    write_header(ws_cf, 1, headers)
    for col_idx in range(1, len(headers) + 1):
        ws_cf.column_dimensions[get_column_letter(col_idx)].width = 15

    for i, cf in enumerate(rental_result.cashflows, 2):
        ws_cf.cell(row=i, column=1, value=cf.year)
        ws_cf.cell(row=i, column=2, value=round(cf.gross_income))
        ws_cf.cell(row=i, column=3, value=round(cf.operating_expenses))
        ws_cf.cell(row=i, column=4, value=round(cf.noi))
        ws_cf.cell(row=i, column=5, value=round(cf.debt_service))
        ws_cf.cell(row=i, column=6, value=round(cf.net_cash_flow))
        ws_cf.cell(row=i, column=7, value=round(cf.cumulative_cf))
        ws_cf.cell(row=i, column=8, value=round(cf.loan_balance))
        ws_cf.cell(row=i, column=9, value=round(cf.property_value))
        for c in range(1, 10):
            ws_cf.cell(row=i, column=c).border = thin_border
            if c >= 2:
                ws_cf.cell(row=i, column=c).number_format = "#,##0"

    # --- 融資比較シート ---
    ws_loan = wb.create_sheet("融資比較")
    loan_headers = ["金融機関", "金利(%)", "融資額", "期間(年)", "月額返済", "年間返済", "総返済額", "総利息", "利用可否", "推奨"]
    write_header(ws_loan, 1, loan_headers)
    for col_idx in range(1, len(loan_headers) + 1):
        ws_loan.column_dimensions[get_column_letter(col_idx)].width = 15

    for i, lr in enumerate(loan_results, 2):
        ws_loan.cell(row=i, column=1, value=lr.bank_name)
        ws_loan.cell(row=i, column=2, value=lr.interest_rate)
        ws_loan.cell(row=i, column=3, value=lr.loan_amount)
        ws_loan.cell(row=i, column=4, value=lr.loan_years)
        ws_loan.cell(row=i, column=5, value=lr.monthly_payment)
        ws_loan.cell(row=i, column=6, value=lr.annual_payment)
        ws_loan.cell(row=i, column=7, value=lr.total_payment)
        ws_loan.cell(row=i, column=8, value=lr.total_interest)
        ws_loan.cell(row=i, column=9, value="○" if lr.available else f"× {lr.reason}")
        ws_loan.cell(row=i, column=10, value="★" if lr.recommended else "")
        for c in range(1, 11):
            ws_loan.cell(row=i, column=c).border = thin_border
            if c in (3, 5, 6, 7, 8):
                ws_loan.cell(row=i, column=c).number_format = "#,##0"

    # --- 出口戦略シート ---
    ws_exit = wb.create_sheet("出口戦略")
    exit_headers = ["売却年", "物件価値", "ローン残高", "累計CF", "売却手取り", "総利益", "年率ROI(%)", "譲渡所得税"]
    write_header(ws_exit, 1, exit_headers)
    for col_idx in range(1, len(exit_headers) + 1):
        ws_exit.column_dimensions[get_column_letter(col_idx)].width = 15

    for i, es in enumerate(exit_scenarios, 2):
        ws_exit.cell(row=i, column=1, value=f"{es.year}年目")
        ws_exit.cell(row=i, column=2, value=es.property_value)
        ws_exit.cell(row=i, column=3, value=es.loan_balance)
        ws_exit.cell(row=i, column=4, value=es.cumulative_cf)
        ws_exit.cell(row=i, column=5, value=es.sale_proceeds)
        ws_exit.cell(row=i, column=6, value=es.total_profit)
        ws_exit.cell(row=i, column=7, value=es.annualized_roi)
        ws_exit.cell(row=i, column=8, value=es.capital_gains_tax)
        for c in range(1, 9):
            ws_exit.cell(row=i, column=c).border = thin_border
            if c in (2, 3, 4, 5, 6, 8):
                ws_exit.cell(row=i, column=c).number_format = "#,##0"

    # --- リスクシート ---
    ws_risk = wb.create_sheet("リスク分析")
    risk_headers = ["分類", "重要度", "リスク内容", "対策"]
    write_header(ws_risk, 1, risk_headers)
    ws_risk.column_dimensions["A"].width = 10
    ws_risk.column_dimensions["B"].width = 10
    ws_risk.column_dimensions["C"].width = 50
    ws_risk.column_dimensions["D"].width = 50

    for i, r in enumerate(risks, 2):
        ws_risk.cell(row=i, column=1, value=r.category)
        ws_risk.cell(row=i, column=2, value=r.severity)
        ws_risk.cell(row=i, column=3, value=r.description)
        ws_risk.cell(row=i, column=4, value=r.mitigation)
        for c in range(1, 5):
            ws_risk.cell(row=i, column=c).border = thin_border
        if r.severity == "高":
            ws_risk.cell(row=i, column=2).fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

    wb.save(str(output_path))
